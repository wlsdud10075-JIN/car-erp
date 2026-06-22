"""
서류 양식 xlsx 의 "유령 열(phantom column)" 잔재 제거 — 1회성 외과 수술.

배경 (2026-06-22):
  한글 정부양식(xlsm/HWP) 변환 잔재로 매입 3종 양식(말소신청서·말소계약서·위임장)이
  256열마다 양식 컬럼 패턴이 반복되는 <col> 정의를 16,384열까지 들고 있었다.
  다운로드 시 PhpSpreadsheet 가 저장하면서 이 유령 열을 전부 실제 열로 펼쳐
  시트 범위가 A1:WVX 로 폭발 → 엑셀이 "손상·복구" 처리 → 복구 중 열너비/서식 리셋 →
  좌우 정렬 흐트러짐. (정상 양식은 유령 범위가 단일 range 라 안 펼쳐짐 → 제외)

방식 (advisor 권고 = 최소 변경):
  zipfile 로 xl/worksheets/sheetN.xml 의 <cols> 블록만 수정한다.
  - 이미지(도장)·하이퍼링크·sharedStrings·스타일·기타 엔트리는 byte 그대로 복사.
  - keepCol = max(값 있는 셀의 최대 열, 병합범위의 최대 열)  ← 양식 실제 폭
  - <col min..max> 중 min > keepCol → 삭제(유령), max > keepCol 인 straddler → max 를 keepCol 로 clip
    (clip 이 핵심: 삭제하면 P/Q 가 기본폭으로 돌아가 양식이 밀림)
  - <dimension ref> 의 열 부분도 keepCol 로 정정 (런타임은 어차피 재계산하나 템플릿 자체도 깔끔하게)

실행: python scripts/strip-phantom-columns.py
git 추적 파일을 인플레이스 수정하므로, 문제 시 git checkout 으로 복원.
"""
import os
import re
import shutil
import zipfile

from openpyxl.utils import column_index_from_string, get_column_letter

TARGETS = [
    "resources/templates/system/deregistration_application.xlsx",
    "resources/templates/system/deregistration_contract.xlsx",
    "resources/templates/system/power_of_attorney.xlsx",
    "resources/templates/karaba/deregistration_application.xlsx",
    "resources/templates/karaba/deregistration_contract.xlsx",
    "resources/templates/karaba/power_of_attorney.xlsx",
]

SHEET_RE = re.compile(r"xl/worksheets/sheet\d+\.xml$")


def clean_sheet_xml(xml: str):
    """sheet XML 1개의 <cols> 를 정리. (새 xml, keepCol or None) 반환."""
    keep = 0
    # 값 있는 셀: <c r="COL#" ...> 바로 뒤에 <v|<f|<is
    for m in re.finditer(r'<c r="([A-Z]+)\d+"[^>]*?>(?=<(?:v|f|is)\b)', xml):
        keep = max(keep, column_index_from_string(m.group(1)))
    # 병합범위 끝 열
    for m in re.finditer(r'<mergeCell ref="[A-Z]+\d+:([A-Z]+)\d+"', xml):
        keep = max(keep, column_index_from_string(m.group(1)))

    if keep == 0:
        return xml, None  # 빈 시트 — 손대지 않음

    changed = False
    cols_m = re.search(r"<cols>(.*?)</cols>", xml, re.S)
    if cols_m:
        new_cols = []
        for col in re.finditer(r"<col [^>]*?/>", cols_m.group(1)):
            tag = col.group(0)
            mn = int(re.search(r'min="(\d+)"', tag).group(1))
            mx = int(re.search(r'max="(\d+)"', tag).group(1))
            if mn > keep:
                changed = True
                continue  # 유령 열 삭제
            if mx > keep:
                tag = re.sub(r'max="\d+"', f'max="{keep}"', tag)  # straddler clip
                changed = True
            new_cols.append(tag)
        if changed:
            replacement = ("<cols>" + "".join(new_cols) + "</cols>") if new_cols else ""
            xml = xml[: cols_m.start()] + replacement + xml[cols_m.end():]

    # <dimension ref="A1:WVX39"/> → A1:Q39
    keepL = get_column_letter(keep)
    xml_new = re.sub(
        r'<dimension ref="([A-Z]+\d+):[A-Z]+(\d+)"\s*/>',
        lambda m: f'<dimension ref="{m.group(1)}:{keepL}{m.group(2)}"/>',
        xml,
    )
    if xml_new != xml:
        xml = xml_new
        changed = True

    return xml, (keep if changed else None)


def process(path: str):
    tmp = path + ".tmp"
    touched = []
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for it in zin.infolist():
                data = zin.read(it.filename)
                if SHEET_RE.search(it.filename):
                    new_xml, keep = clean_sheet_xml(data.decode("utf-8"))
                    if keep:
                        data = new_xml.encode("utf-8")
                        touched.append(f"{it.filename}->keep col{keep}({get_column_letter(keep)})")
                zi = zipfile.ZipInfo(it.filename, date_time=it.date_time)
                zi.compress_type = it.compress_type
                zi.external_attr = it.external_attr
                zi.internal_attr = it.internal_attr
                zi.create_system = it.create_system
                zout.writestr(zi, data)
    shutil.move(tmp, path)
    return touched


if __name__ == "__main__":
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for rel in TARGETS:
        p = os.path.join(root, rel)
        touched = process(p)
        print(f"{rel}: {', '.join(touched) if touched else 'no change'}")
